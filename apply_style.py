# apply_styles.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def apply_styles(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    # Define styles
    orange_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    gray_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    bold_font = Font(bold=True, name="Poppins")
    header_font = Font(bold=True, name="Poppins", size=12)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    # Style for cell A1
    worksheet['A1'].fill = orange_fill
    worksheet['A1'].font = bold_font
    worksheet['A1'].alignment = center_alignment

    # Style for column headers
    for column in worksheet.iter_cols(min_row=1, max_row=1):
        max_width = 0
        for cell in column:
            cell.fill = orange_fill
            cell.font = header_font
            cell.font = bold_font
            cell.value = cell.value.capitalize()
            cell.alignment = center_alignment
            column_letter = get_column_letter(cell.column)
            if len(cell.value) > max_width:
                max_width = len(cell.value)
        
        worksheet.column_dimensions[column_letter].width = max_width + 5

    # Style for cells in the first column (excluding the header)
    for row in worksheet.iter_rows(min_row=2):
        cell = row[0]
        cell.fill = gray_fill
        cell.font = bold_font
        cell.alignment = left_alignment
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        column_letter = get_column_letter(cell.column)
        worksheet.column_dimensions[column_letter].width = max(len(cell.value) + 11, 10)  # Adjust column width

    # Style for cells after the first column
    for column in worksheet.iter_cols(min_col=2):
        for cell in column:
            cell.alignment = center_alignment
            cell.font = Font(name="Poppins")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Additional styles
    red_font = Font(color="FF0000", name="Poppins")
    red_bck_bold_font = Font(color="FF0000", bold=True, name="Poppins")

    # Apply styles to corresponding cells and hide rows without competitor prices
    rows_to_hide = []
    for row in worksheet.iter_rows(min_row=2):
        tienda_solar_price_cell = row[1]
        try:
            tienda_solar_price = float(tienda_solar_price_cell.value)
        except (TypeError, ValueError):
            print(f"Invalid value in cell {tienda_solar_price_cell.coordinate}: {tienda_solar_price_cell.value}")
            continue

        competitor_price_found = False
        for cell in row[2:]:
            if cell.value is not None:
                try:
                    cell_value = float(cell.value)
                    competitor_price_found = True
                except (TypeError, ValueError):
                    print(f"Invalid value in cell {cell.coordinate}: {cell.value}")
                    continue

                if cell_value < tienda_solar_price:
                    tienda_solar_price_cell.fill = red_fill
                    tienda_solar_price_cell.font = bold_font
                    break

        if not competitor_price_found:
            rows_to_hide.append(row[0].row)

    for row_num in rows_to_hide:
        worksheet.row_dimensions[row_num].hidden = True

    # Save changes to the Excel file
    workbook.save(file_path)

if __name__ == "__main__":
    import sys
    file_path = sys.argv[1]
    apply_styles(file_path)
